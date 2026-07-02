from __future__ import annotations

from dataclasses import dataclass
from io import BytesIO
import math
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


POINT_ALIASES = ["point name", "point", "name", "point id", "pt", "ptid", "station", "mark"]
EASTING_ALIASES = ["easting", "east", "e", "x", "grid e", "grid east"]
NORTHING_ALIASES = ["northing", "north", "n", "y", "grid n", "grid north"]
HEIGHT_ALIASES = ["height", "elevation", "elev", "z", "rl", "level", "orthometric height"]
SOLUTION_ALIASES = ["solution", "fix status", "fix", "status", "quality", "gnss solution"]

OUTPUT_COLUMNS = [
    "Point Name",
    "Control Easting",
    "Control Northing",
    "Control Height",
    "Ground Check Easting",
    "Ground Check Northing",
    "Ground Check Height / Observed NGL",
    "Ground Delta E mm",
    "Ground Delta N mm",
    "Ground Displacement mm",
    "Ground Height Difference mm",
    "Ground Position Status",
    "Nail Check Easting",
    "Nail Check Northing",
    "Nail Check Height",
    "Nail Height Difference mm",
    "Nail Position Displacement mm",
    "Nail Height Status",
    "Final Status",
    "Remarks",
]


@dataclass(frozen=True)
class ColumnMap:
    point: str
    easting: str
    northing: str
    height: str
    solution: str

    def as_dict(self) -> dict[str, str]:
        return {
            "point": self.point,
            "easting": self.easting,
            "northing": self.northing,
            "height": self.height,
            "solution": self.solution,
        }


def normalize_name(value: Any) -> str:
    if value is None or pd.isna(value):
        return ""
    return str(value).strip()


def normalize_header(value: Any) -> str:
    text = normalize_name(value).lower()
    return " ".join(text.replace("_", " ").replace("-", " ").split())


def detect_columns(df: pd.DataFrame) -> tuple[dict[str, str | None], list[str]]:
    normalized = {normalize_header(column): column for column in df.columns}
    aliases = {
        "point": POINT_ALIASES,
        "easting": EASTING_ALIASES,
        "northing": NORTHING_ALIASES,
        "height": HEIGHT_ALIASES,
        "solution": SOLUTION_ALIASES,
    }
    detected: dict[str, str | None] = {}
    missing: list[str] = []

    for field, options in aliases.items():
        match = None
        for option in options:
            if option in normalized:
                match = normalized[option]
                break
        if match is None:
            for header, original in normalized.items():
                if any(option in header for option in options if len(option) > 1):
                    match = original
                    break
        detected[field] = match
        if match is None:
            missing.append(field)

    return detected, missing


def load_fieldbook(uploaded_file: Any) -> pd.DataFrame:
    name = getattr(uploaded_file, "name", "").lower()
    if name.endswith((".xlsx", ".xlsm", ".xls")):
        return pd.read_excel(uploaded_file)
    return pd.read_csv(uploaded_file)


def _numeric(series: pd.Series) -> pd.Series:
    return pd.to_numeric(series, errors="coerce")


def _mm(value: float | None) -> float | None:
    if value is None or pd.isna(value):
        return None
    return round(float(value) * 1000.0, 1)


def _coord(value: float | None) -> float | None:
    if value is None or pd.isna(value):
        return None
    return round(float(value), 3)


def _status(pass_flag: bool, pass_text: str, fail_text: str) -> str:
    return pass_text if pass_flag else fail_text


def classify_rows(df: pd.DataFrame, columns: ColumnMap) -> pd.DataFrame:
    classified = df.copy()
    classified["_PointKey"] = classified[columns.point].map(normalize_name)
    classified["_SolutionText"] = classified[columns.solution].map(normalize_name)
    classified["_Class"] = classified["_SolutionText"].str.upper().eq("NONE").map(
        {True: "CONTROL", False: "MEASURED"}
    )
    return classified


def analyze_fieldbook(
    df: pd.DataFrame,
    columns: ColumnMap,
    position_tolerance_mm: float = 100.0,
    height_tolerance_mm: float = 20.0,
    invert_for_autocad: bool = False,
) -> dict[str, pd.DataFrame]:
    classified = classify_rows(df, columns)
    errors: list[dict[str, Any]] = []

    for source, target in [
        (columns.easting, "_EastingNum"),
        (columns.northing, "_NorthingNum"),
        (columns.height, "_HeightNum"),
    ]:
        classified[target] = _numeric(classified[source])

    if invert_for_autocad:
        classified["_EastingReport"] = -classified["_EastingNum"]
        classified["_NorthingReport"] = -classified["_NorthingNum"]
    else:
        classified["_EastingReport"] = classified["_EastingNum"]
        classified["_NorthingReport"] = classified["_NorthingNum"]

    for idx, row in classified.iterrows():
        missing = []
        if not row["_PointKey"]:
            missing.append("point name")
        for label, field in [
            ("easting", "_EastingNum"),
            ("northing", "_NorthingNum"),
            ("height", "_HeightNum"),
        ]:
            if pd.isna(row[field]):
                missing.append(label)
        if missing:
            errors.append(
                {
                    "Input Row": int(idx) + 2,
                    "Point Name": row.get(columns.point, ""),
                    "Issue": "Missing or invalid " + ", ".join(missing),
                }
            )

    controls = classified[classified["_Class"] == "CONTROL"].copy()
    measured = classified[classified["_Class"] == "MEASURED"].copy()
    measured_keys = set(measured["_PointKey"])
    control_keys = set(controls["_PointKey"])
    unmatched = measured[~measured["_PointKey"].isin(control_keys)].copy()

    rows: list[dict[str, Any]] = []
    used_measured_indexes: set[Any] = set()

    for _, control in controls.iterrows():
        point_key = control["_PointKey"]
        if not point_key:
            continue
        candidates = measured[measured["_PointKey"] == point_key].copy()
        usable = candidates.dropna(subset=["_EastingNum", "_NorthingNum", "_HeightNum"])

        if usable.empty:
            rows.append(_empty_result_row(control, "No valid measured observation matched this control point."))
            continue

        usable["_dE"] = usable["_EastingNum"] - control["_EastingNum"]
        usable["_dN"] = usable["_NorthingNum"] - control["_NorthingNum"]
        usable["_dH"] = usable["_HeightNum"] - control["_HeightNum"]
        usable["_disp"] = (usable["_dE"].pow(2) + usable["_dN"].pow(2)).pow(0.5)

        ground = usable.sort_values(["_disp", "_dH"], key=lambda s: s.abs() if s.name == "_dH" else s).iloc[0]
        nail = usable.iloc[usable["_dH"].abs().argsort()].iloc[0]
        used_measured_indexes.update([ground.name, nail.name])

        ground_disp_mm = _mm(ground["_disp"])
        ground_height_diff_mm = _mm(ground["_dH"])
        nail_height_diff_mm = _mm(nail["_dH"])
        nail_disp_mm = _mm(nail["_disp"])

        position_pass = ground_disp_mm is not None and ground_disp_mm <= position_tolerance_mm
        height_pass = (
            nail_height_diff_mm is not None and abs(nail_height_diff_mm) <= height_tolerance_mm
        )
        final_status = "OVERALL PASS" if position_pass and height_pass else "OVERALL FAIL"

        remarks = []
        if len(candidates) > 1:
            remarks.append(f"{len(candidates)} measured observations found.")
        if ground.name == nail.name:
            remarks.append("Same observation used for ground and nail checks.")

        rows.append(
            {
                "Point Name": control["_PointKey"],
                "Control Easting": _coord(control["_EastingReport"]),
                "Control Northing": _coord(control["_NorthingReport"]),
                "Control Height": _coord(control["_HeightNum"]),
                "Ground Check Easting": _coord(ground["_EastingReport"]),
                "Ground Check Northing": _coord(ground["_NorthingReport"]),
                "Ground Check Height / Observed NGL": _coord(ground["_HeightNum"]),
                "Ground Delta E mm": _mm(ground["_dE"]),
                "Ground Delta N mm": _mm(ground["_dN"]),
                "Ground Displacement mm": ground_disp_mm,
                "Ground Height Difference mm": ground_height_diff_mm,
                "Ground Position Status": _status(position_pass, "PASS POSITION", "FAIL POSITION"),
                "Nail Check Easting": _coord(nail["_EastingReport"]),
                "Nail Check Northing": _coord(nail["_NorthingReport"]),
                "Nail Check Height": _coord(nail["_HeightNum"]),
                "Nail Height Difference mm": nail_height_diff_mm,
                "Nail Position Displacement mm": nail_disp_mm,
                "Nail Height Status": _status(height_pass, "PASS HEIGHT", "FAIL HEIGHT"),
                "Final Status": final_status,
                "Remarks": " ".join(remarks),
            }
        )

    raw = classified.copy()
    raw["Classification"] = raw["_Class"]
    raw.loc[raw.index.isin(unmatched.index), "Classification"] = "UNMATCHED"
    raw = raw.drop(columns=[col for col in raw.columns if col.startswith("_")], errors="ignore")

    report = pd.DataFrame(rows, columns=OUTPUT_COLUMNS)
    failed = report[
        (report["Ground Position Status"] == "FAIL POSITION")
        | (report["Nail Height Status"] == "FAIL HEIGHT")
        | (report["Final Status"] == "OVERALL FAIL")
    ].copy()

    summary = _build_summary(report, unmatched, position_tolerance_mm, height_tolerance_mm)

    unmatched_out = unmatched.drop(columns=[col for col in unmatched.columns if col.startswith("_")], errors="ignore")
    unmatched_out["Classification"] = "UNMATCHED"

    return {
        "Report": report,
        "Summary": summary,
        "Failed Points": failed,
        "Unmatched Measurements": unmatched_out,
        "Raw Classified Data": raw,
        "Errors": pd.DataFrame(errors, columns=["Input Row", "Point Name", "Issue"]),
    }


def _empty_result_row(control: pd.Series, remark: str) -> dict[str, Any]:
    row = {column: None for column in OUTPUT_COLUMNS}
    row.update(
        {
            "Point Name": control["_PointKey"],
            "Control Easting": _coord(control["_EastingReport"]),
            "Control Northing": _coord(control["_NorthingReport"]),
            "Control Height": _coord(control["_HeightNum"]),
            "Ground Position Status": "FAIL POSITION",
            "Nail Height Status": "FAIL HEIGHT",
            "Final Status": "OVERALL FAIL",
            "Remarks": remark,
        }
    )
    return row


def _build_summary(
    report: pd.DataFrame,
    unmatched: pd.DataFrame,
    position_tolerance_mm: float,
    height_tolerance_mm: float,
) -> pd.DataFrame:
    total = len(report)
    position_pass = int((report["Ground Position Status"] == "PASS POSITION").sum()) if total else 0
    height_pass = int((report["Nail Height Status"] == "PASS HEIGHT").sum()) if total else 0
    overall_pass = int((report["Final Status"] == "OVERALL PASS").sum()) if total else 0

    def pct(count: int) -> float:
        return round((count / total) * 100, 1) if total else 0.0

    return pd.DataFrame(
        [
            {"Metric": "Total matched control points", "Value": total},
            {"Metric": "Total unmatched measured points", "Value": len(unmatched)},
            {"Metric": "Position tolerance mm", "Value": position_tolerance_mm},
            {"Metric": "Height tolerance mm", "Value": height_tolerance_mm},
            {"Metric": "Position pass count", "Value": position_pass},
            {"Metric": "Position pass percentage", "Value": pct(position_pass)},
            {"Metric": "Height pass count", "Value": height_pass},
            {"Metric": "Height pass percentage", "Value": pct(height_pass)},
            {"Metric": "Overall pass count", "Value": overall_pass},
            {"Metric": "Overall pass percentage", "Value": pct(overall_pass)},
        ]
    )


def export_excel(sheets: dict[str, pd.DataFrame]) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)

    for sheet_name, df in sheets.items():
        ws = wb.create_sheet(title=sheet_name[:31])
        if df.empty:
            ws.append(list(df.columns) if len(df.columns) else ["No records"])
        else:
            ws.append(list(df.columns))
            for values in df.itertuples(index=False, name=None):
                ws.append(list(values))
        _format_sheet(ws, sheet_name)

    output = BytesIO()
    wb.save(output)
    return output.getvalue()


def _format_sheet(ws: Any, sheet_name: str) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    pass_fill = PatternFill("solid", fgColor="C6EFCE")
    fail_fill = PatternFill("solid", fgColor="FFC7CE")
    unmatched_fill = PatternFill("solid", fgColor="FCE4D6")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    if ws.max_row > 1 and ws.max_column > 0:
        table_ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
        table_name = "".join(ch for ch in sheet_name.title() if ch.isalnum())[:20] or "Report"
        table = Table(displayName=f"{table_name}Table", ref=table_ref)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        ws.add_table(table)

    status_columns = {
        cell.value: idx + 1 for idx, cell in enumerate(ws[1]) if isinstance(cell.value, str)
    }
    for row in ws.iter_rows(min_row=2):
        values = [cell.value for cell in row]
        is_unmatched = "UNMATCHED" in values
        for cell in row:
            text = str(cell.value).upper() if cell.value is not None else ""
            if "PASS" in text:
                cell.fill = pass_fill
            elif "FAIL" in text:
                cell.fill = fail_fill
            elif is_unmatched:
                cell.fill = unmatched_fill

    for column_cells in ws.columns:
        max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = min(max(max_length + 2, 12), 36)

