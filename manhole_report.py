from __future__ import annotations

import argparse
import math
import re
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import ScatterChart, Reference, Series
from openpyxl.drawing.image import Image as XlsxImage
from openpyxl.styles import Alignment, Font, PatternFill, Side, Border
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from reportlab.lib import colors
from reportlab.lib.pagesizes import A3, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import Image as PdfImage
from reportlab.platypus import PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table as PdfTable, TableStyle
from reportlab.graphics.shapes import Drawing, Line, String, Circle, Rect


DEFAULT_SOURCE = (
    r"C:\Users\User\OneDrive\Documents\Survey-TR7-1\KESHE\Ashikuni"
    r"\Fransfontein\Survey\fransfontein-SET.csv"
)
DEFAULT_OUTPUT = "outputs/fransfontein_manhole_settingout_report.xlsx"
DEFAULT_PDF_OUTPUT = "outputs/fransfontein_manhole_layout_longsection.pdf"
DEFAULT_DXF_OUTPUT = "outputs/fransfontein_manhole_layout_longsection.dxf"
DEFAULT_LOGO = "assets/kesheshiwe_logo.png"

DESIGN_ZERO_FIELDS = [
    "Antenna Height",
    "SD to Base",
    "HD to Base",
    "Diff Age",
    "HRMS",
    "VRMS",
    "PDOP",
    "SAT Locked",
    "SAT Visible",
    "Input Antenna Height",
    "Average Times",
]

TRENCH_CODES = {"TR", "TREC", "MHINV"}
PEG_WIRE_CODES = {"PEG", "WIRE"}
SPOT_SHOT_CODES = {"SS"}
LINE_BREAK_DISTANCE_M = 100.0


def _num(value: Any) -> float | None:
    try:
        if pd.isna(value):
            return None
        number = float(value)
    except (TypeError, ValueError):
        return None
    return number if math.isfinite(number) else None


def _clean_code(value: Any) -> str:
    if pd.isna(value):
        return ""
    return str(value).strip().upper()


def _round(value: float | None, places: int = 3) -> float | None:
    if value is None or pd.isna(value):
        return None
    return round(float(value), places)


def _mm(value: float | None) -> float | None:
    if value is None or pd.isna(value):
        return None
    return round(float(value) * 1000.0, 0)


def _has_zero_observation_metadata(row: pd.Series) -> bool:
    present = [field for field in DESIGN_ZERO_FIELDS if field in row.index]
    if not present:
        return False
    zeros = 0
    for field in present:
        value = _num(row[field])
        if value is not None and abs(value) < 1e-9:
            zeros += 1
    return zeros >= max(4, int(len(present) * 0.65))


def _valid_coord(row: pd.Series) -> bool:
    north = _num(row.get("North"))
    east = _num(row.get("East"))
    height = _num(row.get("Height"))
    return (
        north is not None
        and east is not None
        and height is not None
        and not (abs(north) < 1e-9 and abs(east) < 1e-9 and abs(height) < 1e-9)
    )


def _measurement_type(code: str) -> str:
    if code in SPOT_SHOT_CODES:
        return "Spot shot"
    if code in PEG_WIRE_CODES:
        return "Peg / Wire"
    if code in TRENCH_CODES:
        return "Trench / invert"
    if code == "BM":
        return "Benchmark"
    return "Other"


def _point_sort_key(name: Any) -> tuple[str, int, str]:
    text = "" if pd.isna(name) else str(name).strip().upper()
    match = re.match(r"([A-Z]+)\D*(\d+)", text)
    if match:
        return (match.group(1), int(match.group(2)), text)
    return (text, 10**9, text)


def _nearest_design(row: pd.Series, design: pd.DataFrame) -> pd.Series | None:
    if design.empty or not _valid_coord(row):
        return None
    d_n = design["NorthNum"] - float(row["NorthNum"])
    d_e = design["EastNum"] - float(row["EastNum"])
    distances = (d_n.pow(2) + d_e.pow(2)).pow(0.5)
    idx = distances.idxmin()
    matched = design.loc[idx].copy()
    matched["MatchDistance"] = float(distances.loc[idx])
    return matched


def _action_from_difference(diff_m: float | None, tolerance_m: float) -> str:
    if diff_m is None:
        return ""
    if abs(diff_m) <= tolerance_m:
        return "On grade"
    return "Cut" if diff_m > 0 else "Fill"


def _build_surveyed_comparison(
    assigned: pd.DataFrame,
    tolerance_m: float,
    max_match_distance_m: float,
) -> pd.DataFrame:
    rows: list[dict[str, Any]] = []
    for _, row in assigned.iterrows():
        level_diff_m = None
        if pd.notna(row.get("Measured Level")) and pd.notna(row.get("Design Level")):
            level_diff_m = float(row["Measured Level"]) - float(row["Design Level"])
        distance_m = row.get("Match Distance m")
        outside_match = pd.notna(distance_m) and float(distance_m) > max_match_distance_m
        rows.append(
            {
                "Survey Point": row.get("Measured Point"),
                "Code": row.get("Code"),
                "Type": row.get("Measurement Type"),
                "Survey East m": row.get("East"),
                "Survey North m": row.get("North"),
                "Survey Level m": row.get("Measured Level"),
                "Nearest Manhole": row.get("Assigned Design Point"),
                "Design Level m": row.get("Design Level"),
                "Distance m": row.get("Match Distance m"),
                "Level Diff m": _round(level_diff_m),
                "Cut m": _round(max(level_diff_m, 0.0)) if level_diff_m is not None else None,
                "Fill m": _round(max(-level_diff_m, 0.0)) if level_diff_m is not None else None,
                "Action": "Check distance" if outside_match else _action_from_difference(level_diff_m, tolerance_m),
                "Remarks": f"Nearest manhole is more than {max_match_distance_m:.0f} m away."
                if outside_match
                else "",
            }
        )
    return pd.DataFrame(rows)


def _line_label(line_no: int) -> str:
    return f"Line {line_no}"


def _sorted_by_point(df: pd.DataFrame, reverse_flow: bool = False) -> pd.DataFrame:
    sorted_df = df.sort_values("Point Name", key=lambda s: s.map(_point_sort_key))
    if reverse_flow:
        sorted_df = sorted_df.iloc[::-1]
    return sorted_df


def get_design_manhole_names(df: pd.DataFrame) -> list[str]:
    work = df.copy()
    for col in ["North", "East", "Height", "Antenna Height"]:
        work[f"{col}Num"] = pd.to_numeric(work[col], errors="coerce")
    work["Is Valid Coordinate"] = work.apply(_valid_coord, axis=1)
    design = work[
        work["Solution"].astype(str).str.upper().eq("NONE")
        & work["Is Valid Coordinate"]
        & work.apply(_has_zero_observation_metadata, axis=1)
    ].copy()
    return (
        design.sort_values("Point Name", key=lambda s: s.map(_point_sort_key))["Point Name"]
        .dropna()
        .astype(str)
        .tolist()
    )


def _apply_flow_order(
    df: pd.DataFrame,
    point_col: str,
    flow_order: list[str] | None = None,
    reverse_flow: bool = False,
) -> pd.DataFrame:
    if df.empty:
        return df
    ordered = df.copy()
    if flow_order:
        order_lookup = {str(name): idx for idx, name in enumerate(flow_order)}
        ordered["_FlowOrder"] = ordered[point_col].astype(str).map(order_lookup)
        ordered = ordered[ordered["_FlowOrder"].notna()].sort_values("_FlowOrder")
        ordered = ordered.drop(columns=["_FlowOrder"])
    else:
        ordered = ordered.sort_values(point_col, key=lambda s: s.map(_point_sort_key))
    if reverse_flow:
        ordered = ordered.iloc[::-1]
    return ordered


def _build_report(
    df: pd.DataFrame,
    level_tolerance_mm: float = 20.0,
    max_match_distance_m: float = 25.0,
    excluded_points: list[str] | None = None,
    reverse_flow: bool = False,
    flow_order: list[str] | None = None,
) -> dict[str, pd.DataFrame]:
    work = df.copy()
    for col in ["North", "East", "Height", "Antenna Height"]:
        work[f"{col}Num"] = pd.to_numeric(work[col], errors="coerce")
    work["CodeClean"] = work.get("Code", "").map(_clean_code)
    work["Measurement Type"] = work["CodeClean"].map(_measurement_type)
    work["Is Valid Coordinate"] = work.apply(_valid_coord, axis=1)
    work["Is Design Point"] = (
        work["Solution"].astype(str).str.upper().eq("NONE")
        & work["Is Valid Coordinate"]
        & work.apply(_has_zero_observation_metadata, axis=1)
    )

    excluded = {str(point).strip() for point in (excluded_points or []) if str(point).strip()}
    design = work[work["Is Design Point"]].copy()
    if excluded:
        design = design[~design["Point Name"].astype(str).str.strip().isin(excluded)].copy()
    if flow_order:
        selected = {str(point).strip() for point in flow_order if str(point).strip()}
        design = design[design["Point Name"].astype(str).str.strip().isin(selected)].copy()
    measured = work[~work["Is Design Point"] & work["Is Valid Coordinate"]].copy()

    assigned_rows: list[dict[str, Any]] = []
    for idx, row in measured.iterrows():
        nearest = _nearest_design(row, design)
        if nearest is None:
            assigned_name = None
            match_distance = None
            design_level = None
        else:
            assigned_name = nearest["Point Name"]
            match_distance = nearest["MatchDistance"]
            design_level = nearest["HeightNum"]
        assigned_rows.append(
            {
                "Input Row": int(idx) + 2,
                "Measured Point": row.get("Point Name"),
                "Code": row.get("Code"),
                "Measurement Type": row["Measurement Type"],
                "Local Time": row.get("Local Time"),
                "North": _round(row["NorthNum"]),
                "East": _round(row["EastNum"]),
                "Measured Level": _round(row["HeightNum"]),
                "Assigned Design Point": assigned_name,
                "Match Distance m": _round(match_distance),
                "Design Level": _round(design_level),
                "Level Difference mm": _mm(row["HeightNum"] - design_level)
                if design_level is not None
                else None,
                "Action": _action_from_difference(
                    (row["HeightNum"] - design_level) if design_level is not None else None,
                    level_tolerance_mm / 1000.0,
                ),
                "HRMS": row.get("HRMS"),
                "VRMS": row.get("VRMS"),
                "Solution": row.get("Solution"),
            }
        )
    assigned = pd.DataFrame(assigned_rows)
    surveyed_comparison = _build_surveyed_comparison(
        assigned,
        level_tolerance_mm / 1000.0,
        max_match_distance_m,
    )

    comparison_rows: list[dict[str, Any]] = []
    tolerance_m = level_tolerance_mm / 1000.0
    for _, design_row in _apply_flow_order(
        design,
        "Point Name",
        flow_order=flow_order,
        reverse_flow=reverse_flow,
    ).iterrows():
        point = design_row["Point Name"]
        nearby = assigned[
            (assigned["Assigned Design Point"] == point)
            & (assigned["Match Distance m"].fillna(max_match_distance_m + 1) <= max_match_distance_m)
        ].copy()

        def pick_nearest(types: set[str]) -> pd.Series | None:
            subset = nearby[nearby["Measurement Type"].isin(types)]
            if subset.empty:
                return None
            return subset.sort_values("Match Distance m").iloc[0]

        def pick_lowest(types: set[str]) -> pd.Series | None:
            subset = nearby[nearby["Measurement Type"].isin(types)]
            if subset.empty:
                return None
            return subset.sort_values(["Measured Level", "Match Distance m"]).iloc[0]

        ss = pick_nearest({"Spot shot"})
        peg_wire = pick_nearest({"Peg / Wire"})
        trench = pick_lowest({"Trench / invert"})

        trench_diff = None if trench is None else float(trench["Measured Level"]) - float(design_row["HeightNum"])
        if nearby.empty:
            remarks = "No measured points within match distance."
        elif trench is None:
            remarks = "No trench/invert point within match distance."
        else:
            remarks = ""
        comparison_rows.append(
            {
                "Manhole": point,
                "Design East m": _round(design_row["EastNum"]),
                "Design North m": _round(design_row["NorthNum"]),
                "Design Level m": _round(design_row["HeightNum"]),
                "Survey Point": None if trench is None else trench["Measured Point"],
                "Survey Code": None if trench is None else trench["Code"],
                "Survey East m": None if trench is None else trench["East"],
                "Survey North m": None if trench is None else trench["North"],
                "Survey Level m": None if trench is None else trench["Measured Level"],
                "Level Diff m": _round(trench_diff),
                "Cut m": _round(max(trench_diff, 0.0)) if trench_diff is not None else None,
                "Fill m": _round(max(-trench_diff, 0.0)) if trench_diff is not None else None,
                "Action": _action_from_difference(trench_diff, tolerance_m),
                "Remarks": remarks,
            }
        )

    comparison = pd.DataFrame(comparison_rows)

    trench_obs = assigned[
        assigned["Measurement Type"].eq("Trench / invert")
        & assigned["Assigned Design Point"].notna()
        & (assigned["Match Distance m"].fillna(max_match_distance_m + 1) <= max_match_distance_m)
    ].copy()
    trench_obs = trench_obs.sort_values(
        ["Assigned Design Point", "Measured Level"], key=lambda s: s.map(_point_sort_key) if s.name == "Assigned Design Point" else s
    )
    lowest_trench = trench_obs.sort_values(["Assigned Design Point", "Measured Level"]).drop_duplicates(
        "Assigned Design Point", keep="first"
    )
    design_lookup = design.set_index("Point Name")
    slope_rows: list[dict[str, Any]] = []
    last: pd.Series | None = None
    chainage = 0.0
    line_no = 1
    lowest_trench = _apply_flow_order(
        lowest_trench,
        "Assigned Design Point",
        flow_order=flow_order,
        reverse_flow=reverse_flow,
    )
    for _, row in lowest_trench.iterrows():
        design_point = row["Assigned Design Point"]
        current_design = design_lookup.loc[design_point]
        if last is None:
            run = None
            rise = None
            slope = None
            note = "Start"
        else:
            run = math.hypot(float(row["North"]) - float(last["North"]), float(row["East"]) - float(last["East"]))
            rise = float(row["Measured Level"]) - float(last["Measured Level"])
            if run and run > LINE_BREAK_DISTANCE_M:
                line_no += 1
                chainage = 0.0
                run = None
                rise = None
                slope = None
                note = f"New line: previous manhole is more than {LINE_BREAK_DISTANCE_M:.0f} m away"
            else:
                slope = (rise / run) * 100.0 if run else None
                note = ""
                chainage += run or 0.0
        slope_rows.append(
            {
                "Line": _line_label(line_no),
                "Design Point": design_point,
                "Chainage m": round(chainage, 3),
                "Measured Trench Point": row["Measured Point"],
                "North": row["North"],
                "East": row["East"],
                "Design Level": _round(current_design["HeightNum"]),
                "Measured Trench Level": row["Measured Level"],
                "Trench Diff mm": row["Level Difference mm"],
                "Run from Previous m": _round(run),
                "Rise from Previous m": _round(rise),
                "Current Slope %": _round(slope, 3),
                "Note": note,
            }
        )
        last = row
    slope = pd.DataFrame(slope_rows)

    qa = work[
        (~work["Is Design Point"])
        & (
            ~work["Is Valid Coordinate"]
            | ((work["Antenna HeightNum"].fillna(0).abs() < 1e-9) & work["Solution"].astype(str).str.upper().ne("NONE"))
        )
    ][
        [
            "Point Name",
            "Code",
            "North",
            "East",
            "Height",
            "Antenna Height",
            "Solution",
            "Local Time",
            "HRMS",
            "VRMS",
            "Point Type",
        ]
    ].copy()

    summary = pd.DataFrame(
        [
            ["Source rows", len(df)],
            ["Design points", len(design)],
            ["Valid measured observations", len(measured)],
            ["Design points with trench/invert", int(comparison["Survey Point"].notna().sum())],
            ["Design points needing cut", int((comparison["Action"] == "Cut").sum())],
            ["Design points needing fill", int((comparison["Action"] == "Fill").sum())],
            ["Level tolerance mm", level_tolerance_mm],
            ["Maximum match distance m", max_match_distance_m],
            ["Line break distance m", LINE_BREAK_DISTANCE_M],
            ["Flow direction", "Reverse selected flow" if reverse_flow else "Selected flow"],
            ["Selected flow", " -> ".join(flow_order or []) if flow_order else "Point-name order"],
            ["Excluded design points", ", ".join(sorted(excluded)) if excluded else "None"],
            ["Classification rule", "Design = Solution NONE + valid coordinates + mostly zero observation metadata"],
            ["Trench rule", "Lowest TR / tr / trec / MHINV level associated to the design point"],
        ],
        columns=["Metric", "Value"],
    )

    raw_cols = [
        "Point Name",
        "Code",
        "North",
        "East",
        "Height",
        "Antenna Height",
        "Solution",
        "Local Time",
        "Diff Age",
        "HRMS",
        "VRMS",
        "Point Type",
        "Measurement Type",
        "Is Design Point",
        "Is Valid Coordinate",
    ]
    raw = work[[col for col in raw_cols if col in work.columns]].copy()

    return {
        "Summary": summary,
        "Manhole Comparison": comparison,
        "Surveyed Comparison": surveyed_comparison,
        "Trench Slopes": slope,
        "Measured Assignments": assigned,
        "QA Excluded Rows": qa,
        "Raw Classified Data": raw,
    }


def _write_df(ws: Any, df: pd.DataFrame, start_row: int = 6) -> None:
    while ws.max_row < start_row - 1:
        ws.append([None])
    ws.append(list(df.columns))
    for row in df.itertuples(index=False, name=None):
        ws.append([None if pd.isna(value) else value for value in row])


def _add_excel_header(ws: Any, title: str, logo_path: str | Path | None = None) -> None:
    ws.merge_cells("B1:H1")
    ws.merge_cells("B2:H2")
    ws.merge_cells("B3:H3")
    ws["B1"] = "Kesheshiwe Engineering Surveyors Cc"
    ws["B2"] = title
    ws["B3"] = "Manhole setting-out report - all measurements in metres"
    ws["B1"].font = Font(bold=True, size=14, color="17324D")
    ws["B2"].font = Font(bold=True, size=11)
    ws["B3"].font = Font(size=9, color="4A5568")
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 18
    ws.row_dimensions[3].height = 18
    if logo_path and Path(logo_path).exists():
        image = XlsxImage(str(logo_path))
        image.width = 86
        image.height = 86
        ws.add_image(image, "A1")


def _add_category_headers(ws: Any, table_name: str) -> None:
    if table_name == "Manhole Comparison":
        categories = [
            ("A5:A5", "Manhole"),
            ("B5:D5", "Design Coordinates"),
            ("E5:I5", "Surveyed Coordinates"),
            ("J5:M5", "Comparison"),
            ("N5:N5", "Notes"),
        ]
    elif table_name == "Surveyed Comparison":
        categories = [
            ("A5:C5", "Surveyed Point"),
            ("D5:F5", "Surveyed Coordinates"),
            ("G5:H5", "Nearest Manhole"),
            ("I5:M5", "Comparison"),
        ]
    else:
        return
    fill = PatternFill("solid", fgColor="D9EAF7")
    for cell_range, label in categories:
        if ":" in cell_range and cell_range.split(":")[0] != cell_range.split(":")[1]:
            ws.merge_cells(cell_range)
        cell = ws[cell_range.split(":")[0]]
        cell.value = label
        cell.fill = fill
        cell.font = Font(bold=True, color="17324D")
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _format_sheet(ws: Any, table_name: str, header_row: int = 6) -> None:
    dark = PatternFill("solid", fgColor="17324D")
    pale = PatternFill("solid", fgColor="EAF2F8")
    cut = PatternFill("solid", fgColor="F8CBAD")
    fill = PatternFill("solid", fgColor="BDD7EE")
    on_grade = PatternFill("solid", fgColor="C6E0B4")
    border = Border(bottom=Side(style="thin", color="D9E2EC"))

    _add_category_headers(ws, table_name)
    ws.freeze_panes = f"A{header_row + 1}"
    ws.sheet_view.showGridLines = False
    ws.page_setup.paperSize = ws.PAPERSIZE_A3
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    two_cm_inches = 2.0 / 2.54
    ws.page_margins.left = two_cm_inches
    ws.page_margins.right = two_cm_inches
    ws.page_margins.top = two_cm_inches
    ws.page_margins.bottom = two_cm_inches
    data_col_count = max(
        (idx for idx, cell in enumerate(ws[header_row], start=1) if cell.value is not None),
        default=1,
    )
    data_last_col = get_column_letter(data_col_count)
    ws.print_title_rows = f"1:{header_row}"
    ws.auto_filter.ref = f"A{header_row}:{data_last_col}{ws.max_row}"
    for cell in ws[header_row]:
        cell.fill = dark
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[header_row].height = 32
    for row in ws.iter_rows(min_row=header_row + 1):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=False)
            text = str(cell.value).upper() if cell.value is not None else ""
            if text == "CUT":
                cell.fill = cut
            elif text == "FILL":
                cell.fill = fill
            elif text == "ON GRADE":
                cell.fill = on_grade
        if row[0].row % 2 == 0:
            for cell in row:
                if cell.fill.fill_type is None:
                    cell.fill = pale

    if ws.max_row > 1 and ws.max_column > 1:
        safe_name = re.sub(r"[^A-Za-z0-9]", "", table_name)[:20] or "Data"
        ref = f"A{header_row}:{data_last_col}{ws.max_row}"
        table = Table(displayName=f"{safe_name}Table", ref=ref)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        ws.add_table(table)

    width_profiles = {
        "Manhole Comparison": [9, 11, 12, 11, 12, 9, 11, 12, 11, 10, 8, 8, 9, 22],
        "Surveyed Comparison": [12, 8, 13, 11, 12, 11, 12, 11, 10, 10, 8, 8, 12, 22],
        "Trench Slopes": [9, 11, 10, 14, 12, 11, 11, 12, 10, 10, 10, 10, 22],
        "Summary": [26, 52],
    }
    profile = width_profiles.get(table_name)
    for column_cells in ws.columns:
        col_idx = column_cells[0].column
        if profile and col_idx <= len(profile):
            width = profile[col_idx - 1]
        else:
            max_len = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
            width = min(max(max_len + 2, 9), 18)
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def _add_slope_chart(ws: Any, header_row: int = 6) -> None:
    if ws.max_row < header_row + 2:
        return
    chart = ScatterChart()
    chart.title = "Measured trench profile"
    chart.style = 13
    chart.x_axis.title = "Chainage (m)"
    chart.y_axis.title = "Level (m)"
    chart.height = 9
    chart.width = 18

    xvalues = Reference(ws, min_col=3, min_row=header_row + 1, max_row=ws.max_row)
    measured = Reference(ws, min_col=8, min_row=header_row + 1, max_row=ws.max_row)
    design = Reference(ws, min_col=7, min_row=header_row + 1, max_row=ws.max_row)
    series_measured = Series(measured, xvalues, title="Measured trench")
    series_design = Series(design, xvalues, title="Design level")
    chart.series.append(series_measured)
    chart.series.append(series_design)
    ws.add_chart(chart, "N2")


def export_report(
    sheets: dict[str, pd.DataFrame],
    output_path: str | Path,
    logo_path: str | Path | None = DEFAULT_LOGO,
) -> Path:
    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    wb.remove(wb.active)
    for sheet_name, df in sheets.items():
        ws = wb.create_sheet(sheet_name[:31])
        _add_excel_header(ws, sheet_name, logo_path=logo_path)
        _write_df(ws, df)
        _format_sheet(ws, sheet_name)
        if sheet_name == "Trench Slopes":
            _add_slope_chart(ws)

    wb.save(output)
    return output


def _bounds(values: list[tuple[float, float]]) -> tuple[float, float, float, float]:
    xs = [x for x, _ in values]
    ys = [y for _, y in values]
    return min(xs), min(ys), max(xs), max(ys)


def _scale_points(
    points: list[tuple[float, float]],
    x: float,
    y: float,
    width: float,
    height: float,
) -> list[tuple[float, float]]:
    min_x, min_y, max_x, max_y = _bounds(points)
    span_x = max(max_x - min_x, 1.0)
    span_y = max(max_y - min_y, 1.0)
    scale = min(width / span_x, height / span_y)
    used_w = span_x * scale
    used_h = span_y * scale
    off_x = x + (width - used_w) / 2.0
    off_y = y + (height - used_h) / 2.0
    return [(off_x + (px - min_x) * scale, off_y + (py - min_y) * scale) for px, py in points]


def _plot_coord(east: Any, north: Any) -> tuple[float, float]:
    return -float(east), -float(north)


def _label_offset(index: int) -> tuple[float, float]:
    offsets = [(6, 6), (6, -12), (-18, 6), (-18, -12), (10, 0), (-22, 0)]
    return offsets[index % len(offsets)]


def _header(title: str, subtitle: str, logo_path: str | Path | None = DEFAULT_LOGO) -> PdfTable:
    if logo_path and Path(logo_path).exists():
        logo: Any = PdfImage(str(logo_path), width=55, height=55)
    else:
        logo = PdfTable(
            [[Paragraph("<b>KES</b>", getSampleStyleSheet()["Title"])]],
            colWidths=[55],
            rowHeights=[38],
        )
        logo.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#17324D")),
                    ("TEXTCOLOR", (0, 0), (-1, -1), colors.white),
                    ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ]
            )
        )
    styles = getSampleStyleSheet()
    text = Paragraph(
        f"<b>Kesheshiwe Engineering Surveyors Cc</b><br/>{title}<br/><font size='8'>{subtitle}</font>",
        styles["Normal"],
    )
    table = PdfTable([[logo, text]], colWidths=[65, 720])
    table.setStyle(
        TableStyle(
            [
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
                ("LINEBELOW", (0, 0), (-1, -1), 1, colors.HexColor("#17324D")),
            ]
        )
    )
    return table


def _layout_drawing(slope: pd.DataFrame, comparison: pd.DataFrame, width: float = 760, height: float = 420) -> Drawing:
    drawing = Drawing(width, height)
    drawing.add(Rect(0, 0, width, height, strokeColor=colors.HexColor("#B8C4CE"), fillColor=None))
    if slope.empty:
        drawing.add(String(20, height / 2, "No trench/invert points available for layout.", fontSize=10))
        return drawing

    points = [_plot_coord(row["East"], row["North"]) for _, row in slope.iterrows()]
    scaled = _scale_points(points, 35, 35, width - 70, height - 70)
    rows = list(slope.reset_index(drop=True).iterrows())
    for line_name, group in slope.groupby("Line", sort=False):
        indexes = list(group.index)
        line_points = [scaled[list(slope.index).index(idx)] for idx in indexes]
        for first, second in zip(line_points, line_points[1:]):
            drawing.add(Line(first[0], first[1], second[0], second[1], strokeColor=colors.HexColor("#0F766E"), strokeWidth=1.4))
    action_lookup = comparison.set_index("Manhole")["Action"].to_dict()
    fill_lookup = {
        "Cut": colors.HexColor("#D9480F"),
        "Fill": colors.HexColor("#2563EB"),
        "On grade": colors.HexColor("#2F855A"),
        None: colors.HexColor("#718096"),
        "": colors.HexColor("#718096"),
    }
    for label_index, ((idx, row), (px, py)) in enumerate(zip(rows, scaled)):
        point = row["Design Point"]
        action = action_lookup.get(point)
        drawing.add(Circle(px, py, 4, fillColor=fill_lookup.get(action, colors.HexColor("#718096")), strokeColor=colors.black, strokeWidth=0.4))
        dx, dy = _label_offset(label_index)
        drawing.add(String(px + dx, py + dy, str(point), fontSize=6, fillColor=colors.black))
    drawing.add(String(20, height - 18, "Layout: selected flow, plotted with East/North multiplied by -1", fontSize=9))
    drawing.add(String(20, 14, "Units: metres. Tables keep source coordinates; layout plot uses inverted coordinates for orientation.", fontSize=7, fillColor=colors.HexColor("#4A5568")))
    return drawing


def _longsection_drawing(line_df: pd.DataFrame, width: float = 760, height: float = 300) -> Drawing:
    drawing = Drawing(width, height)
    drawing.add(Rect(0, 0, width, height, strokeColor=colors.HexColor("#B8C4CE"), fillColor=None))
    if line_df.empty:
        return drawing
    left, bottom, plot_w, plot_h = 55, 45, width - 95, height - 90
    chainages = [float(v) for v in line_df["Chainage m"]]
    levels = [float(v) for v in line_df["Measured Trench Level"]] + [float(v) for v in line_df["Design Level"]]
    min_c, max_c = min(chainages), max(chainages)
    min_l, max_l = min(levels), max(levels)
    if max_c - min_c < 1:
        max_c = min_c + 1
    pad_l = max((max_l - min_l) * 0.15, 0.5)
    min_l -= pad_l
    max_l += pad_l

    def xy(chain: float, level: float) -> tuple[float, float]:
        return (
            left + ((chain - min_c) / (max_c - min_c)) * plot_w,
            bottom + ((level - min_l) / (max_l - min_l)) * plot_h,
        )

    drawing.add(Line(left, bottom, left + plot_w, bottom, strokeColor=colors.black, strokeWidth=0.8))
    drawing.add(Line(left, bottom, left, bottom + plot_h, strokeColor=colors.black, strokeWidth=0.8))
    measured_pts = [xy(float(row["Chainage m"]), float(row["Measured Trench Level"])) for _, row in line_df.iterrows()]
    design_pts = [xy(float(row["Chainage m"]), float(row["Design Level"])) for _, row in line_df.iterrows()]
    for first, second in zip(measured_pts, measured_pts[1:]):
        drawing.add(Line(first[0], first[1], second[0], second[1], strokeColor=colors.HexColor("#2563EB"), strokeWidth=1.3))
    for first, second in zip(design_pts, design_pts[1:]):
        drawing.add(Line(first[0], first[1], second[0], second[1], strokeColor=colors.HexColor("#D9480F"), strokeWidth=1.0))
    for (_, row), (mx, my), (dx, dy) in zip(line_df.iterrows(), measured_pts, design_pts):
        drawing.add(Circle(mx, my, 3, fillColor=colors.HexColor("#2563EB"), strokeColor=None))
        drawing.add(Circle(dx, dy, 2.5, fillColor=colors.HexColor("#D9480F"), strokeColor=None))
        drawing.add(String(mx + 3, bottom - 16, str(row["Design Point"]), fontSize=6))
        slope = row["Current Slope %"]
        if pd.notna(slope):
            drawing.add(String(mx - 8, my + 7, f"{float(slope):.2f}%", fontSize=6, fillColor=colors.HexColor("#2D3748")))
    drawing.add(String(20, height - 18, f"{line_df.iloc[0]['Line']} longsection", fontSize=9))
    drawing.add(String(left + plot_w - 145, height - 18, "Blue = measured trench, Orange = design", fontSize=7))
    drawing.add(String(18, bottom + plot_h - 8, "Level (m)", fontSize=7))
    drawing.add(String(left + (plot_w / 2) - 24, bottom - 34, "Chainage (m)", fontSize=7))
    drawing.add(String(left, bottom - 30, f"0 m", fontSize=7))
    drawing.add(String(left + plot_w - 42, bottom - 30, f"{max_c:.1f} m", fontSize=7))
    drawing.add(String(18, bottom, f"{min_l:.2f}", fontSize=7))
    drawing.add(String(18, bottom + plot_h - 4, f"{max_l:.2f}", fontSize=7))
    return drawing


def export_pdf(
    sheets: dict[str, pd.DataFrame],
    output_path: str | Path,
    logo_path: str | Path | None = DEFAULT_LOGO,
) -> Path:
    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)
    doc = SimpleDocTemplate(
        str(output),
        pagesize=landscape(A3),
        leftMargin=28,
        rightMargin=28,
        topMargin=24,
        bottomMargin=24,
    )
    styles = getSampleStyleSheet()
    story: list[Any] = []
    summary = sheets["Summary"]
    comparison = sheets["Manhole Comparison"]
    slope = sheets["Trench Slopes"]
    story.append(_header("Manhole Setting-Out Field Report", "Layout and longsection - printable A3, all measurements in metres", logo_path=logo_path))
    story.append(Spacer(1, 8))
    metrics = [[str(a), str(b)] for a, b in summary.itertuples(index=False, name=None)]
    metrics_table = PdfTable([["Metric", "Value"]] + metrics, colWidths=[190, 560])
    metrics_table.setStyle(_pdf_table_style())
    story.append(metrics_table)
    story.append(Spacer(1, 12))
    story.append(_layout_drawing(slope, comparison))
    story.append(Spacer(1, 12))

    printable_cols = [
        "Manhole",
        "Design East m",
        "Design North m",
        "Design Level m",
        "Survey Point",
        "Survey Code",
        "Survey East m",
        "Survey North m",
        "Survey Level m",
        "Level Diff m",
        "Cut m",
        "Fill m",
        "Action",
    ]
    preview = comparison[printable_cols].copy()
    table_rows = [list(preview.columns)] + [
        ["" if pd.isna(v) else v for v in row] for row in preview.itertuples(index=False, name=None)
    ]
    report_table = PdfTable(table_rows, repeatRows=1, colWidths=[55, 70, 75, 65, 70, 50, 70, 75, 65, 60, 45, 45, 55])
    report_table.setStyle(_pdf_table_style())
    story.append(Paragraph("<b>Field Comparison Summary</b>", styles["Heading3"]))
    story.append(report_table)

    for _, line_df in slope.groupby("Line", sort=False):
        story.append(PageBreak())
        story.append(_header("Manhole Longsection", f"{line_df.iloc[0]['Line']} - new line starts where spacing exceeds 100 m", logo_path=logo_path))
        story.append(Spacer(1, 8))
        story.append(_longsection_drawing(line_df))
        story.append(Spacer(1, 8))
        cols = ["Design Point", "Chainage m", "Design Level", "Measured Trench Level", "Run from Previous m", "Current Slope %", "Note"]
        line_table = line_df[cols].copy()
        line_rows = [cols] + [["" if pd.isna(v) else v for v in row] for row in line_table.itertuples(index=False, name=None)]
        lt = PdfTable(line_rows, repeatRows=1, colWidths=[85, 75, 85, 110, 105, 90, 245])
        lt.setStyle(_pdf_table_style())
        story.append(lt)
        story.append(Spacer(1, 16))

    doc.build(story, onFirstPage=_page_footer, onLaterPages=_page_footer)
    return output


def _pdf_table_style() -> TableStyle:
    return TableStyle(
        [
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#17324D")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 7),
            ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#CBD5E0")),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F7FAFC")]),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("ALIGN", (1, 1), (-1, -1), "RIGHT"),
        ]
    )


def _page_footer(canvas: Any, doc: Any) -> None:
    canvas.saveState()
    canvas.setFont("Helvetica", 7)
    canvas.setFillColor(colors.HexColor("#4A5568"))
    canvas.drawString(28, 14, "Kesheshiwe Engineering Surveyors Cc - Manhole setting-out report")
    canvas.drawRightString(doc.pagesize[0] - 28, 14, f"Page {doc.page}")
    canvas.restoreState()


def _dxf_pair(code: int, value: Any) -> str:
    return f"{code}\n{value}\n"


def _dxf_line(layer: str, x1: float, y1: float, x2: float, y2: float) -> str:
    return (
        _dxf_pair(0, "LINE")
        + _dxf_pair(8, layer)
        + _dxf_pair(10, x1)
        + _dxf_pair(20, y1)
        + _dxf_pair(30, 0)
        + _dxf_pair(11, x2)
        + _dxf_pair(21, y2)
        + _dxf_pair(31, 0)
    )


def _dxf_circle(layer: str, x: float, y: float, radius: float) -> str:
    return (
        _dxf_pair(0, "CIRCLE")
        + _dxf_pair(8, layer)
        + _dxf_pair(10, x)
        + _dxf_pair(20, y)
        + _dxf_pair(30, 0)
        + _dxf_pair(40, radius)
    )


def _dxf_text(layer: str, x: float, y: float, text: str, height: float = 1.5) -> str:
    return (
        _dxf_pair(0, "TEXT")
        + _dxf_pair(8, layer)
        + _dxf_pair(10, x)
        + _dxf_pair(20, y)
        + _dxf_pair(30, 0)
        + _dxf_pair(40, height)
        + _dxf_pair(1, text)
    )


def export_dxf(sheets: dict[str, pd.DataFrame], output_path: str | Path) -> Path:
    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)
    slope = sheets["Trench Slopes"]
    comparison = sheets["Manhole Comparison"]
    action_lookup = comparison.set_index("Manhole")["Action"].to_dict()
    parts = [_dxf_pair(0, "SECTION"), _dxf_pair(2, "ENTITIES")]
    for line_name, group in slope.groupby("Line", sort=False):
        coords = [_plot_coord(r["East"], r["North"]) for _, r in group.iterrows()]
        for first, second in zip(coords, coords[1:]):
            parts.append(_dxf_line(f"{line_name}_LAYOUT", first[0], first[1], second[0], second[1]))
        for label_index, ((_, row), (east, north)) in enumerate(zip(group.iterrows(), coords)):
            point = str(row["Design Point"])
            layer = f"{line_name}_{action_lookup.get(point) or 'NO_ACTION'}".replace(" ", "_").upper()
            parts.append(_dxf_circle(layer, east, north, 1.25))
            dx, dy = _label_offset(label_index)
            parts.append(_dxf_text("MANHOLE_LABELS", east + dx * 0.35, north + dy * 0.35, point, 1.4))

    y_offset = 0.0
    for line_name, group in slope.groupby("Line", sort=False):
        base_y = y_offset
        last = None
        for _, row in group.iterrows():
            x = float(row["Chainage m"])
            measured_y = base_y + float(row["Measured Trench Level"])
            design_y = base_y + float(row["Design Level"])
            parts.append(_dxf_circle(f"{line_name}_LONGSECTION_MEASURED", x, measured_y, 0.25))
            parts.append(_dxf_circle(f"{line_name}_LONGSECTION_DESIGN", x, design_y, 0.2))
            parts.append(_dxf_text("LONGSECTION_LABELS", x, measured_y + 0.6, str(row["Design Point"]), 0.8))
            if last is not None:
                parts.append(_dxf_line(f"{line_name}_LONGSECTION_MEASURED", last[0], last[1], x, measured_y))
                parts.append(_dxf_line(f"{line_name}_LONGSECTION_DESIGN", last[0], last[2], x, design_y))
            last = (x, measured_y, design_y)
        y_offset += 35.0

    parts.extend([_dxf_pair(0, "ENDSEC"), _dxf_pair(0, "EOF")])
    output.write_text("".join(parts), encoding="ascii")
    return output


def main() -> None:
    parser = argparse.ArgumentParser(description="Generate a manhole setting-out report workbook.")
    parser.add_argument("--source", default=DEFAULT_SOURCE, help="Input CSV or Excel fieldbook.")
    parser.add_argument("--output", default=DEFAULT_OUTPUT, help="Output .xlsx path.")
    parser.add_argument("--pdf-output", default=DEFAULT_PDF_OUTPUT, help="Output printable PDF path.")
    parser.add_argument("--dxf-output", default=DEFAULT_DXF_OUTPUT, help="Output CAD DXF path.")
    parser.add_argument("--logo", default=DEFAULT_LOGO, help="Company logo image path.")
    parser.add_argument("--reverse-flow", action="store_true", help="Reverse the manhole flow direction.")
    parser.add_argument(
        "--exclude",
        action="append",
        default=[],
        help="Design manhole to exclude. Repeat for multiple manholes.",
    )
    parser.add_argument(
        "--flow",
        action="append",
        default=[],
        help="Design manhole flow order. Repeat in the desired order.",
    )
    parser.add_argument("--no-pdf-cad", action="store_true", help="Only generate the Excel workbook.")
    parser.add_argument("--level-tolerance-mm", type=float, default=20.0)
    parser.add_argument("--max-match-distance-m", type=float, default=25.0)
    args = parser.parse_args()

    source = Path(args.source)
    if source.suffix.lower() in {".xlsx", ".xlsm", ".xls"}:
        df = pd.read_excel(source)
    else:
        df = pd.read_csv(source)

    sheets = _build_report(
        df,
        level_tolerance_mm=args.level_tolerance_mm,
        max_match_distance_m=args.max_match_distance_m,
        excluded_points=args.exclude,
        reverse_flow=args.reverse_flow,
        flow_order=args.flow or None,
    )
    output = export_report(sheets, args.output, logo_path=args.logo)
    print(output.resolve())
    if not args.no_pdf_cad:
        pdf_output = export_pdf(sheets, args.pdf_output, logo_path=args.logo)
        dxf_output = export_dxf(sheets, args.dxf_output)
        print(pdf_output.resolve())
        print(dxf_output.resolve())


if __name__ == "__main__":
    main()
